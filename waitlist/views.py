import datetime
import re
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string, get_template
from django.utils import timezone
from django.utils.html import escape
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail, EmailMultiAlternatives
from xhtml2pdf import pisa
import openpyxl
from openpyxl.utils import get_column_letter
from decouple import config
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from .models import Passenger, PassengerLog
from django.shortcuts import render


# ---------------- Views ----------------



def logs_deleted_success(request):
    return render(request, 'waitlist/logs_deleted.html')




CLEAR_LOGS_PASSWORD = "admin123"  # 🔐 Troque isso por uma senha mais segura

@require_POST
def clear_logs_by_day(request):
    password = request.POST.get("password")
    if password != CLEAR_LOGS_PASSWORD:
        return HttpResponse("❌ Incorrect password", status=403)

    date_str = request.POST.get("date")
    try:
        date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        start = timezone.make_aware(datetime.datetime.combine(date_obj, datetime.time.min))
        end = timezone.make_aware(datetime.datetime.combine(date_obj, datetime.time.max))

        deleted, _ = PassengerLog.objects.filter(time_called__range=(start, end)).delete()
        return redirect('logs_deleted_success')

    except Exception as e:
        return HttpResponse(f"❌ Error: {e}", status=500) and redirect('passenger_logs')

@require_POST
def clear_all_logs(request):
    password = request.POST.get("password")
    if password != CLEAR_LOGS_PASSWORD:
        return HttpResponse("❌ Incorrect password", status=403) and redirect('passenger_logs')

    deleted, _ = PassengerLog.objects.all().delete()
    return redirect('logs_deleted_success')






def cancel_confirmation(request):
    return render(request, 'cancel_confirmation.html')

def cancel_spot(request, guest_id):
    guest = get_object_or_404(Passenger, id=guest_id)
    guest.delete()
    return HttpResponse("✅ Your spot has been successfully cancelled.")



def add_passenger(request):
    if request.method == 'POST':
        name = escape(request.POST['name'])
        country_code = request.POST['country_code']
        local_phone = re.sub(r'\D', '', request.POST['local_phone'])
        guests = int(request.POST.get('guests', 0))
        email = request.POST.get('email', '').strip() or None

        passenger = Passenger.objects.create(
            name=name,
            country_code=country_code,
            local_phone=local_phone,
            guests=guests,
            notification_method='email',
            email=email
        )

        

        if passenger.email:
            cancel_url = request.build_absolute_uri(f"/cancel/{passenger.id}/")
            subject = "The Liffey Lounge Waiting List Confirmation"
            from_email = config('HOST_USER')
            to_email = passenger.email
            context = {'name': passenger.name, 'guests': passenger.guests, 'cancel_url': cancel_url}
            html_content = render_to_string('email/added_to_queue.html', context)
            msg = EmailMultiAlternatives(subject, '', from_email, [to_email])
            msg.attach_alternative(html_content, "text/html")
            msg.send()

        return redirect('list')

    return render(request, 'waitlist/add.html')

def call_specific_passenger(request, id):
    passenger = get_object_or_404(Passenger, id=id)
    passenger.called = True
    passenger.called_at = timezone.now()
    passenger.save()

    PassengerLog.objects.create(
        name=passenger.name,
        country_code=passenger.country_code,
        phone=passenger.local_phone,
        guests=passenger.guests,
        time_joined=passenger.joined_at
    )

    if passenger.email:
        subject = "Lounge Spot Ready"
        from_email = config('HOST_USER')
        to_email = passenger.email
        context = {'name': passenger.name}
        try:
            html_content = render_to_string('email/passenger_called.html', context)
            msg = EmailMultiAlternatives(subject, '', from_email, [to_email])
            msg.attach_alternative(html_content, "text/html")
            msg.send()
        except Exception:
            send_mail(subject, from_email, [to_email])

    return redirect('list')

def call_next_passenger(request):
    if request.method == 'POST':
        next_passenger = Passenger.objects.filter(called=False).order_by('joined_at').first()
        if next_passenger:
            return call_specific_passenger(request, next_passenger.id)
    return redirect('list')

def auto_remove_passenger(request, id):
    passenger = get_object_or_404(Passenger, id=id)
    if passenger.called and passenger.called_at:
        elapsed = timezone.now() - passenger.called_at
        if elapsed.total_seconds() >= 600:
            passenger.delete()
    

def delete_passenger(request, pk):
    passenger = get_object_or_404(Passenger, pk=pk)
    passenger.delete()
    return redirect('list')


def passenger_list(request):
    passengers = Passenger.objects.order_by('joined_at')
    return render(request, 'waitlist/list.html', {'passengers': passengers})

def passenger_detail(request, id):
    passenger = get_object_or_404(Passenger, id=id)
    return render(request, 'waitlist/passenger_detail.html', {'passenger': passenger})

def passenger_logs(request):
    logs = PassengerLog.objects.order_by('-time_called')
    return render(request, 'waitlist/logs.html', {'logs': logs})

def export_logs_pdf(request):
    logs = PassengerLog.objects.order_by('-time_called')
    template_path = 'waitlist/logs_pdf.html'
    context = {'logs': logs}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="passenger_logs.pdf"'
    html = get_template(template_path).render(context)
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    return response

def export_logs_excel(request):
    logs = PassengerLog.objects.order_by('-time_called')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Passenger Logs"
    headers = ['Name', 'Phone', 'Guests', 'Time Joined', 'Time Called']
    ws.append(headers)

    for log in logs:
        ws.append([
            log.name,
            f"+{log.country_code}{log.phone}",
            log.guests,
            log.time_joined.strftime("%Y-%m-%d %H:%M"),
            log.time_called.strftime("%Y-%m-%d %H:%M") if log.time_called else ''
        ])

    for i, column in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=passenger_logs.xlsx'
    wb.save(response)
    return response
